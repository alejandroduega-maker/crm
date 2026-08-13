"""Aplicación CRM de telefonía. Punto de entrada WSGI."""
from __future__ import annotations

import io
import os
import time
import traceback
from datetime import date, datetime

from jinja2 import Environment, FileSystemLoader, select_autoescape

from . import db, negocio
from .auth import (comprobar, cifrar, crear_usuario, password_temporal, puede_ver_cliente,
                   requiere_admin, requiere_login, usuario_actual, usuario_por_credenciales,
                   validar_password)
from .framework import (Request, Response, Router, SessionCookie, json_response, redirect,
                        servir_estatico, token_csrf, STATUS)

BASE = os.path.dirname(os.path.abspath(__file__))
SECRETO = os.environ.get("CRM_SECRETO", "cambia-esta-clave-en-produccion-por-favor")
sesiones = SessionCookie(SECRETO)
rutas = Router()

# ------------------------------------------------------------------ plantillas
jinja = Environment(
    loader=FileSystemLoader(os.path.join(BASE, "templates")),
    autoescape=select_autoescape(["html"]),
    trim_blocks=True, lstrip_blocks=True,
)


def eur(valor):
    if valor in (None, ""):
        return "—"
    return f"{float(valor):,.2f} €".replace(",", "@").replace(".", ",").replace("@", ".")


def eur0(valor):
    if valor in (None, ""):
        return "—"
    return f"{float(valor):,.0f} €".replace(",", ".")


def num(valor, dec=0):
    if valor in (None, ""):
        return "—"
    return f"{float(valor):,.{dec}f}".replace(",", "@").replace(".", ",").replace("@", ".")


def fecha(valor, formato="%d/%m/%Y"):
    d = negocio.a_fecha(valor)
    return d.strftime(formato) if d else "—"


def momento(valor):
    try:
        return datetime.fromisoformat(valor).strftime("%d/%m/%Y %H:%M")
    except Exception:
        return valor or "—"


jinja.filters.update(eur=eur, eur0=eur0, num=num, fecha=fecha, momento=momento)
jinja.globals.update(hoy=date.today, ETIQUETAS=db.ETIQUETAS, negocio=negocio)


def render(peticion, plantilla, **ctx):
    ctx.setdefault("usuario", peticion.usuario)
    ctx.setdefault("ruta", peticion.path)
    ctx.setdefault("csrf", peticion.session.get("csrf", ""))
    ctx.setdefault("aviso", peticion.query.get("aviso"))
    ctx.setdefault("aviso_tipo", peticion.query.get("t", "ok"))
    ctx.setdefault("ajustes", peticion.ajustes)
    return Response(jinja.get_template(plantilla).render(**ctx))


# --------------------------------------------------------------------- ayudas
def clientes_visibles(con, usuario, incluir_borrados=False, comercial_id=None):
    sql = "SELECT c.*, u.nombre AS comercial FROM clientes c JOIN usuarios u ON u.id = c.comercial_id WHERE 1=1"
    args = []
    if not incluir_borrados:
        sql += " AND c.borrado = 0"
    if usuario["rol"] != "admin":
        sql += " AND c.comercial_id = ?"
        args.append(usuario["id"])
    elif comercial_id:
        sql += " AND c.comercial_id = ?"
        args.append(comercial_id)
    sql += " ORDER BY c.nombre COLLATE NOCASE"
    u = db.umbrales(con)
    return [negocio.calcular(dict(f), umbrales=u) for f in con.execute(sql, args)]


def cliente_o_none(con, cid, usuario, incluir_borrados=False):
    sql = "SELECT c.*, u.nombre AS comercial FROM clientes c JOIN usuarios u ON u.id=c.comercial_id WHERE c.id=?"
    fila = con.execute(sql, (cid,)).fetchone()
    if not fila:
        return None
    d = dict(fila)
    if d["borrado"] and not incluir_borrados:
        return None
    if not puede_ver_cliente(usuario, d):
        return None
    return negocio.calcular(d, umbrales=db.umbrales(con))


def comerciales(con):
    return [dict(f) for f in con.execute(
        "SELECT id, nombre, usuario, rol, activo FROM usuarios WHERE activo=1 ORDER BY nombre")]


def datos_formulario(peticion, usuario):
    def f(campo, defecto=""):
        return peticion.form.get(campo, defecto).strip()

    def numero(campo, entero=False):
        bruto = f(campo).replace(",", ".")
        if not bruto:
            return 0
        try:
            return int(float(bruto)) if entero else float(bruto)
        except ValueError:
            return 0

    comercial = usuario["id"]
    if usuario["rol"] == "admin" and f("comercial_id"):
        comercial = int(f("comercial_id"))
    return {
        "nombre": f("nombre"), "cif": f("cif"), "persona": f("persona"),
        "telefono": f("telefono"), "email": f("email"), "operador": f("operador"),
        "producto": f("producto"), "num_lineas": numero("num_lineas", True),
        "cuota_linea": numero("cuota_linea"), "fecha_alta": f("fecha_alta") or None,
        "permanencia_meses": numero("permanencia_meses", True),
        "penalizacion_total": numero("penalizacion_total"),
        "estado": f("estado") or "Activo",
        "proxima_accion": f("proxima_accion") or None, "observaciones": f("observaciones"),
        "comercial_id": comercial,
    }


# ---------------------------------------------------------------------- login
@rutas.route("/login", ("GET", "POST"))
def vista_login(peticion):
    if peticion.method == "GET":
        if peticion.usuario:
            return redirect("/")
        return render(peticion, "login.html", error=None)
    con = peticion.con
    usuario = usuario_por_credenciales(con, peticion.get("usuario"), peticion.get("password"))
    if not usuario:
        db.registrar(con, None, "login_fallido", "usuario", None, peticion.get("usuario"), None, peticion.ip)
        con.commit()
        time.sleep(0.4)
        return render(peticion, "login.html", error="Usuario o contraseña incorrectos.")
    con.execute("UPDATE usuarios SET ultimo_acceso=? WHERE id=?", (db.ahora(), usuario["id"]))
    db.registrar(con, usuario, "acceso", "usuario", usuario["id"], usuario["nombre"], None, peticion.ip)
    con.commit()
    peticion.session = {"uid": usuario["id"], "csrf": token_csrf(), "_exp": time.time() + 43200}
    peticion.guardar_sesion = True
    return redirect("/cambiar-password" if usuario["cambiar_password"] else "/")


@rutas.route("/logout")
def vista_logout(peticion):
    if peticion.usuario:
        db.registrar(peticion.con, peticion.usuario, "salida", "usuario",
                     peticion.usuario["id"], peticion.usuario["nombre"], None, peticion.ip)
        peticion.con.commit()
    peticion.session = {}
    peticion.guardar_sesion = True
    return redirect("/login", "Sesión cerrada.")


@rutas.route("/cambiar-password", ("GET", "POST"))
def vista_cambiar_password(peticion):
    if not peticion.usuario:
        return redirect("/login")
    if peticion.method == "GET":
        return render(peticion, "cambiar_password.html", error=None)
    actual, nueva, repetir = peticion.get("actual"), peticion.get("nueva"), peticion.get("repetir")
    if not comprobar(actual, peticion.usuario["password_hash"]):
        return render(peticion, "cambiar_password.html", error="La contraseña actual no es correcta.")
    if nueva != repetir:
        return render(peticion, "cambiar_password.html", error="Las dos contraseñas nuevas no coinciden.")
    problema = validar_password(nueva)
    if problema:
        return render(peticion, "cambiar_password.html", error=problema)
    con = peticion.con
    con.execute("UPDATE usuarios SET password_hash=?, cambiar_password=0 WHERE id=?",
                (cifrar(nueva), peticion.usuario["id"]))
    db.registrar(con, peticion.usuario, "cambio_password", "usuario", peticion.usuario["id"],
                 peticion.usuario["nombre"], None, peticion.ip)
    con.commit()
    return redirect("/", "Contraseña actualizada.")


# -------------------------------------------------------------------- cuadro
@rutas.route("/")
@requiere_login
def vista_panel(peticion):
    con = peticion.con
    filtro_comercial = peticion.query.get("comercial")
    filtro_comercial = int(filtro_comercial) if (filtro_comercial or "").isdigit() else None
    lista = clientes_visibles(con, peticion.usuario, comercial_id=filtro_comercial)
    datos = negocio.resumen(lista)
    uid_tareas = filtro_comercial or (None if peticion.usuario["rol"] == "admin" else peticion.usuario["id"])
    sql_tareas = """
        SELECT t.*, c.nombre AS cliente_nombre 
        FROM tareas_cliente t 
        JOIN clientes c ON c.id = t.cliente_id 
        WHERE t.completada = 0 AND c.borrado = 0
    """
    args_tareas = []
    if uid_tareas:
        sql_tareas += " AND c.comercial_id = ?"
        args_tareas.append(uid_tareas)
    sql_tareas += " ORDER BY t.fecha ASC, t.id DESC LIMIT 6"
    tareas_pendientes = [dict(f) for f in con.execute(sql_tareas, args_tareas)]

    urgentes = sorted([c for c in lista if c["aviso"] in ("urgente", "sin_permanencia")],
                      key=lambda c: c["dias_restantes"])

    return render(
        peticion, "panel.html",
        r=datos, clientes=lista, urgentes=urgentes[:8],
        por_mes=negocio.vencimientos_por_mes(lista),
        por_producto=negocio.agrupar(lista, "producto", lambda c: int(c.get("num_lineas") or 0)),
        comerciales=comerciales(con) if peticion.usuario["rol"] == "admin" else [],
        filtro_comercial=filtro_comercial,
        tareas_pendientes=tareas_pendientes,
    )


# ------------------------------------------------------------------ clientes
@rutas.route("/clientes")
@requiere_login
def vista_clientes(peticion):
    con = peticion.con
    q = peticion.query
    filtro_comercial = int(q["comercial"]) if (q.get("comercial") or "").isdigit() else None
    lista = clientes_visibles(con, peticion.usuario, comercial_id=filtro_comercial)
    texto = (q.get("q") or "").strip().lower()
    if texto:
        lista = [c for c in lista if texto in " ".join(
            str(c.get(k) or "") for k in ("nombre", "cif", "persona", "telefono", "email", "observaciones")).lower()]
    for campo, clave in (("aviso", "aviso"), ("estado", "estado"),
                         ("operador", "operador"), ("prioridad", "prioridad")):
        if q.get(campo):
            lista = [c for c in lista if str(c.get(clave)) == q[campo]]
    orden = q.get("orden", "nombre")
    reverso = orden in ("cuota_total", "penalizacion_pendiente", "puntuacion")
    lista.sort(key=lambda c: (c.get(orden) is None, c.get(orden) if not isinstance(c.get(orden), str)
                              else c[orden].lower()), reverse=reverso)
    return render(peticion, "clientes.html", clientes=lista, q=q,
                  comerciales=comerciales(con) if peticion.usuario["rol"] == "admin" else [],
                  filtro_comercial=filtro_comercial, total=len(lista))


@rutas.route("/vencimientos")
@requiere_login
def vista_vencimientos(peticion):
    con = peticion.con
    filtro_comercial = peticion.query.get("comercial")
    filtro_comercial = int(filtro_comercial) if (filtro_comercial or "").isdigit() else None
    lista = [c for c in clientes_visibles(con, peticion.usuario, comercial_id=filtro_comercial)
             if c["estado"] != "Baja" and c["dias_restantes"] is not None]
    lista.sort(key=lambda c: c["dias_restantes"])

    uid_tareas = filtro_comercial or (None if peticion.usuario["rol"] == "admin" else peticion.usuario["id"])
    sql_tareas = """
        SELECT t.*, c.nombre AS cliente_nombre, u.nombre AS comercial 
        FROM tareas_cliente t 
        JOIN clientes c ON c.id = t.cliente_id 
        JOIN usuarios u ON u.id = c.comercial_id
        WHERE c.borrado = 0
    """
    args_tareas = []
    if uid_tareas:
        sql_tareas += " AND c.comercial_id = ?"
        args_tareas.append(uid_tareas)
    sql_tareas += " ORDER BY t.completada ASC, t.fecha ASC, t.id DESC LIMIT 100"
    tareas_planificadas = [dict(f) for f in con.execute(sql_tareas, args_tareas)]

    return render(peticion, "vencimientos.html", clientes=lista,
                  comerciales=comerciales(con) if peticion.usuario["rol"] == "admin" else [],
                  filtro_comercial=filtro_comercial,
                  tareas_planificadas=tareas_planificadas)


@rutas.route("/clientes/nuevo", ("GET", "POST"))
@requiere_login
def vista_nuevo_cliente(peticion):
    con = peticion.con
    if peticion.method == "GET":
        return render(peticion, "cliente_form.html", cliente={}, comerciales=comerciales(con),
                      titulo="Nuevo cliente", error=None)
    datos = datos_formulario(peticion, peticion.usuario)
    if not datos["nombre"]:
        return render(peticion, "cliente_form.html", cliente=datos, comerciales=comerciales(con),
                      titulo="Nuevo cliente", error="El nombre del cliente es obligatorio.")
    campos = [c for c in db.CAMPOS_CLIENTE]
    con.execute(
        f"INSERT INTO clientes ({','.join(campos)}, creado_en, creado_por, actualizado_en) "
        f"VALUES ({','.join('?' * len(campos))},?,?,?)",
        [datos[c] for c in campos] + [db.ahora(), peticion.usuario["id"], db.ahora()])
    nuevo = con.execute("SELECT last_insert_rowid()").fetchone()[0]
    db.registrar(con, peticion.usuario, "crear", "cliente", nuevo, datos["nombre"],
                 db.diferencias({}, datos), peticion.ip)
    con.commit()
    return redirect(f"/clientes/{nuevo}", "Cliente creado.")


@rutas.route("/clientes/<int:cid>")
@requiere_login
def vista_cliente(peticion, cid):
    con = peticion.con
    cliente = cliente_o_none(con, cid, peticion.usuario)
    if not cliente:
        return render(peticion, "error.html", codigo=404,
                      mensaje="Ese cliente no existe o no es de tu cartera."), 404
    historial = [dict(f) for f in con.execute(
        "SELECT * FROM auditoria WHERE entidad='cliente' AND entidad_id=? ORDER BY momento DESC LIMIT 30", (cid,))]
    tareas = [dict(f) for f in con.execute(
        "SELECT * FROM tareas_cliente WHERE cliente_id=? ORDER BY completada ASC, fecha ASC, id DESC", (cid,))]
    return render(peticion, "cliente_detalle.html", c=cliente, historial=historial, tareas=tareas)


@rutas.route("/clientes/<int:cid>/editar", ("GET", "POST"))
@requiere_login
def vista_editar_cliente(peticion, cid):
    con = peticion.con
    actual = cliente_o_none(con, cid, peticion.usuario)
    if not actual:
        return render(peticion, "error.html", codigo=404, mensaje="Ese cliente no existe o no es de tu cartera.")
    if peticion.method == "GET":
        return render(peticion, "cliente_form.html", cliente=actual, comerciales=comerciales(con),
                      titulo=f"Editar · {actual['nombre']}", error=None)
    datos = datos_formulario(peticion, peticion.usuario)
    if not datos["nombre"]:
        return render(peticion, "cliente_form.html", cliente={**actual, **datos},
                      comerciales=comerciales(con), titulo="Editar", error="El nombre es obligatorio.")
    cambios = db.diferencias(actual, datos)
    campos = list(db.CAMPOS_CLIENTE)
    con.execute(f"UPDATE clientes SET {','.join(c + '=?' for c in campos)}, actualizado_en=? WHERE id=?",
                [datos[c] for c in campos] + [db.ahora(), cid])
    if cambios:
        db.registrar(con, peticion.usuario, "editar", "cliente", cid, datos["nombre"], cambios, peticion.ip)
    con.commit()
    return redirect(f"/clientes/{cid}", "Cambios guardados." if cambios else "No había cambios que guardar.")


@rutas.route("/clientes/<int:cid>/borrar", ("POST",))
@requiere_login
def vista_borrar_cliente(peticion, cid):
    con = peticion.con
    cliente = cliente_o_none(con, cid, peticion.usuario)
    if not cliente:
        return redirect("/clientes", "No se ha encontrado el cliente.", "error")
    con.execute("UPDATE clientes SET borrado=1, borrado_en=?, borrado_por=? WHERE id=?",
                (db.ahora(), peticion.usuario["id"], cid))
    db.registrar(con, peticion.usuario, "borrar", "cliente", cid, cliente["nombre"],
                 {"copia": {k: cliente.get(k) for k in db.CAMPOS_CLIENTE}}, peticion.ip)
    con.commit()
    return redirect("/clientes", f"«{cliente['nombre']}» se ha movido a la papelera. Un administrador puede restaurarlo.")


@rutas.route("/clientes/<int:cid>/tareas/nueva", ("POST",))
@requiere_login
def vista_nueva_tarea(peticion, cid):
    con = peticion.con
    cliente = cliente_o_none(con, cid, peticion.usuario)
    if not cliente:
        return redirect("/clientes", "Cliente no encontrado o no pertenece a tu cartera.", "error")

    fecha = peticion.get("fecha")
    tipo = peticion.get("tipo")
    nota = peticion.get("nota")

    if not fecha or not nota:
        return redirect(f"/clientes/{cid}", "La fecha y la nota son obligatorias.", "error")

    con.execute(
        """INSERT INTO tareas_cliente (cliente_id, fecha, tipo, nota, completada, creado_en, creado_por)
           VALUES (?, ?, ?, ?, 0, ?, ?)""",
        (cid, fecha, tipo, nota, db.ahora(), peticion.usuario["id"])
    )

    db.registrar(con, peticion.usuario, "crear_tarea", "cliente", cid, cliente["nombre"],
                 {"fecha": fecha, "tipo": tipo, "nota": nota}, peticion.ip)
    con.commit()
    return redirect(f"/clientes/{cid}", "Planificación añadida.")


@rutas.route("/clientes/<int:cid>/tareas/<int:tid>/completar", ("POST",))
@requiere_login
def vista_completar_tarea(peticion, cid, tid):
    con = peticion.con
    cliente = cliente_o_none(con, cid, peticion.usuario)
    if not cliente:
        return redirect("/clientes", "Cliente no encontrado o no pertenece a tu cartera.", "error")

    tarea = con.execute("SELECT * FROM tareas_cliente WHERE id=? AND cliente_id=?", (tid, cid)).fetchone()
    if not tarea:
        return redirect(f"/clientes/{cid}", "Planificación no encontrada.", "error")

    con.execute("UPDATE tareas_cliente SET completada=1 WHERE id=?", (tid,))
    db.registrar(con, peticion.usuario, "completar_tarea", "cliente", cid, cliente["nombre"],
                 {"tarea_id": tid, "nota": tarea["nota"]}, peticion.ip)
    con.commit()
    return redirect(f"/clientes/{cid}", "Tarea completada.")


@rutas.route("/clientes/<int:cid>/tareas/<int:tid>/eliminar", ("POST",))
@requiere_login
def vista_eliminar_tarea(peticion, cid, tid):
    con = peticion.con
    cliente = cliente_o_none(con, cid, peticion.usuario)
    if not cliente:
        return redirect("/clientes", "Cliente no encontrado o no pertenece a tu cartera.", "error")

    tarea = con.execute("SELECT * FROM tareas_cliente WHERE id=? AND cliente_id=?", (tid, cid)).fetchone()
    if not tarea:
        return redirect(f"/clientes/{cid}", "Planificación no encontrada.", "error")

    con.execute("DELETE FROM tareas_cliente WHERE id=?", (tid,))
    db.registrar(con, peticion.usuario, "eliminar_tarea", "cliente", cid, cliente["nombre"],
                 {"tarea_id": tid, "nota": tarea["nota"]}, peticion.ip)
    con.commit()
    return redirect(f"/clientes/{cid}", "Planificación eliminada.")


# ------------------------------------------------------- panel de administrador
@rutas.route("/admin/usuarios")
@requiere_admin
def vista_usuarios(peticion):
    con = peticion.con
    filas = [dict(f) for f in con.execute("""
        SELECT u.*, (SELECT COUNT(*) FROM clientes c WHERE c.comercial_id=u.id AND c.borrado=0) AS n_clientes,
               (SELECT COALESCE(SUM(c.num_lineas),0) FROM clientes c
                  WHERE c.comercial_id=u.id AND c.borrado=0 AND c.estado='Activo') AS n_lineas
        FROM usuarios u ORDER BY u.rol, u.nombre""")]
    return render(peticion, "admin_usuarios.html", usuarios=filas,
                  generada=peticion.query.get("nueva"), generada_para=peticion.query.get("para"))


@rutas.route("/admin/usuarios/nuevo", ("POST",))
@requiere_admin
def vista_crear_usuario(peticion):
    con = peticion.con
    usuario = peticion.get("usuario")
    nombre = peticion.get("nombre")
    if not usuario or not nombre:
        return redirect("/admin/usuarios", "Usuario y nombre son obligatorios.", "error")
    if con.execute("SELECT 1 FROM usuarios WHERE usuario=? COLLATE NOCASE", (usuario,)).fetchone():
        return redirect("/admin/usuarios", f"El usuario «{usuario}» ya existe.", "error")
    password = peticion.get("password") or password_temporal()
    problema = validar_password(password)
    if problema:
        return redirect("/admin/usuarios", problema, "error")
    crear_usuario(con, {"usuario": usuario, "nombre": nombre, "email": peticion.get("email"),
                        "rol": peticion.get("rol") or "comercial"},
                  password, peticion.usuario, peticion.ip)
    con.commit()
    return redirect(f"/admin/usuarios?nueva={password}&para={usuario}",
                    "Usuario creado. Apunta la contraseña temporal: solo se muestra ahora.")


@rutas.route("/admin/usuarios/<int:uid>/editar", ("POST",))
@requiere_admin
def vista_editar_usuario(peticion, uid):
    con = peticion.con
    fila = con.execute("SELECT * FROM usuarios WHERE id=?", (uid,)).fetchone()
    if not fila:
        return redirect("/admin/usuarios", "Usuario no encontrado.", "error")
    antes = dict(fila)
    nombre = peticion.get("nombre") or antes["nombre"]
    email = peticion.get("email")
    rol = peticion.get("rol") or antes["rol"]
    activo = 1 if peticion.get("activo") == "1" else 0
    if antes["rol"] == "admin" and rol != "admin":
        otros = con.execute("SELECT COUNT(*) FROM usuarios WHERE rol='admin' AND activo=1 AND id<>?", (uid,)).fetchone()[0]
        if not otros:
            return redirect("/admin/usuarios", "No puedes dejar el sistema sin ningún administrador.", "error")
    if uid == peticion.usuario["id"] and not activo:
        return redirect("/admin/usuarios", "No puedes desactivar tu propia cuenta.", "error")
    con.execute("UPDATE usuarios SET nombre=?, email=?, rol=?, activo=? WHERE id=?",
                (nombre, email, rol, activo, uid))
    cambios = db.diferencias(antes, {"nombre": nombre, "email": email, "rol": rol, "activo": activo},
                             ["nombre", "email", "rol", "activo"])
    if cambios:
        db.registrar(con, peticion.usuario, "editar_usuario", "usuario", uid, nombre, cambios, peticion.ip)
    con.commit()
    return redirect("/admin/usuarios", "Usuario actualizado.")


@rutas.route("/admin/usuarios/<int:uid>/password", ("POST",))
@requiere_admin
def vista_reset_password(peticion, uid):
    con = peticion.con
    fila = con.execute("SELECT * FROM usuarios WHERE id=?", (uid,)).fetchone()
    if not fila:
        return redirect("/admin/usuarios", "Usuario no encontrado.", "error")
    nueva = password_temporal()
    con.execute("UPDATE usuarios SET password_hash=?, cambiar_password=1 WHERE id=?", (cifrar(nueva), uid))
    db.registrar(con, peticion.usuario, "reset_password", "usuario", uid, fila["nombre"], None, peticion.ip)
    con.commit()
    return redirect(f"/admin/usuarios?nueva={nueva}&para={fila['usuario']}",
                    "Contraseña nueva generada. Se la tendrá que cambiar al entrar.")


@rutas.route("/admin/historial")
@requiere_admin
def vista_historial(peticion):
    con = peticion.con
    q = peticion.query
    sql = "SELECT * FROM auditoria WHERE 1=1"
    args = []
    if q.get("accion"):
        sql += " AND accion=?"; args.append(q["accion"])
    if q.get("usuario"):
        sql += " AND usuario_id=?"; args.append(q["usuario"])
    if q.get("texto"):
        sql += " AND (entidad_nombre LIKE ? OR usuario_nombre LIKE ?)"
        args += [f"%{q['texto']}%"] * 2
    sql += " ORDER BY momento DESC, id DESC LIMIT 400"
    registros = []
    for f in con.execute(sql, args):
        d = dict(f)
        try:
            d["cambios_dict"] = __import__("json").loads(d["cambios"]) if d["cambios"] else None
        except Exception:
            d["cambios_dict"] = None
        registros.append(d)
    acciones = [f[0] for f in con.execute("SELECT DISTINCT accion FROM auditoria ORDER BY accion")]
    return render(peticion, "admin_historial.html", registros=registros, acciones=acciones,
                  usuarios=comerciales(con), q=q)


@rutas.route("/admin/papelera")
@requiere_admin
def vista_papelera(peticion):
    con = peticion.con
    filas = [dict(f) for f in con.execute("""
        SELECT c.*, u.nombre AS comercial, b.nombre AS borrado_por_nombre
        FROM clientes c JOIN usuarios u ON u.id=c.comercial_id
        LEFT JOIN usuarios b ON b.id=c.borrado_por
        WHERE c.borrado=1 ORDER BY c.borrado_en DESC""")]
    return render(peticion, "admin_papelera.html", clientes=filas)


@rutas.route("/admin/papelera/<int:cid>/restaurar", ("POST",))
@requiere_admin
def vista_restaurar(peticion, cid):
    con = peticion.con
    fila = con.execute("SELECT * FROM clientes WHERE id=? AND borrado=1", (cid,)).fetchone()
    if not fila:
        return redirect("/admin/papelera", "Ese registro ya no está en la papelera.", "error")
    con.execute("UPDATE clientes SET borrado=0, borrado_en=NULL, borrado_por=NULL WHERE id=?", (cid,))
    db.registrar(con, peticion.usuario, "restaurar", "cliente", cid, fila["nombre"], None, peticion.ip)
    con.commit()
    return redirect("/admin/papelera", f"«{fila['nombre']}» restaurado y visible otra vez para su comercial.")


@rutas.route("/admin/ajustes", ("GET", "POST"))
@requiere_admin
def vista_ajustes(peticion):
    con = peticion.con
    if peticion.method == "POST":
        antes = db.leer_ajustes(con)
        nuevos = {}
        for clave in ("aviso_rojo", "aviso_naranja", "aviso_amarillo"):
            valor = peticion.get(clave)
            if valor.isdigit():
                nuevos[clave] = valor
        if peticion.get("nombre_empresa"):
            nuevos["nombre_empresa"] = peticion.get("nombre_empresa")
        for clave, valor in nuevos.items():
            con.execute("UPDATE ajustes SET valor=? WHERE clave=?", (valor, clave))
        cambios = db.diferencias(antes, nuevos, list(nuevos))
        if cambios:
            db.registrar(con, peticion.usuario, "ajustes", "sistema", None, "Parámetros", cambios, peticion.ip)
        con.commit()
        return redirect("/admin/ajustes", "Parámetros guardados.")
    return render(peticion, "admin_ajustes.html", a=db.leer_ajustes(con))


# -------------------------------------------------------- importar / exportar
@rutas.route("/admin/importar", ("GET", "POST"))
@requiere_admin
def vista_importar(peticion):
    con = peticion.con
    if peticion.method == "GET":
        return render(peticion, "admin_importar.html", comerciales=comerciales(con), resultado=None)
    if "fichero" not in peticion.files:
        return render(peticion, "admin_importar.html", comerciales=comerciales(con),
                      resultado={"error": "No has seleccionado ningún archivo."})
    destino = int(peticion.get("comercial_id") or peticion.usuario["id"])
    nombre_fichero, contenido = peticion.files["fichero"]
    try:
        insertados, saltados = importar_excel(con, contenido, destino, peticion.usuario, peticion.ip)
    except Exception as exc:  # noqa: BLE001
        return render(peticion, "admin_importar.html", comerciales=comerciales(con),
                      resultado={"error": f"No se ha podido leer el archivo: {exc}"})
    con.commit()
    return render(peticion, "admin_importar.html", comerciales=comerciales(con),
                  resultado={"insertados": insertados, "saltados": saltados, "fichero": nombre_fichero})


def importar_excel(con, contenido: bytes, comercial_id: int, autor, ip="-"):
    """Lee la hoja CRM del Excel y da de alta los clientes que no existan."""
    from openpyxl import load_workbook
    libro = load_workbook(io.BytesIO(contenido), data_only=True)
    hoja = libro["CRM"] if "CRM" in libro.sheetnames else libro[libro.sheetnames[0]]
    columnas = {"nombre": 2, "cif": 3, "persona": 4, "telefono": 5, "email": 6, "operador": 7,
                "producto": 8, "num_lineas": 9, "cuota_linea": 10, "fecha_alta": 12,
                "permanencia_meses": 13, "penalizacion_total": 17, "estado": 20,
                "proxima_accion": 24, "observaciones": 26}
    insertados = saltados = 0
    for fila in range(8, hoja.max_row + 1):
        nombre = hoja.cell(row=fila, column=2).value
        if not nombre or not str(nombre).strip():
            continue
        datos = {}
        for campo, col in columnas.items():
            valor = hoja.cell(row=fila, column=col).value
            if campo in ("fecha_alta", "proxima_accion"):
                f = negocio.a_fecha(valor)
                datos[campo] = f.isoformat() if f else None
            elif campo in ("num_lineas", "permanencia_meses"):
                datos[campo] = int(valor or 0)
            elif campo in ("cuota_linea", "penalizacion_total"):
                datos[campo] = float(valor or 0)
            else:
                datos[campo] = str(valor).strip() if valor is not None else ""
        datos["comercial_id"] = comercial_id
        existe = con.execute("SELECT 1 FROM clientes WHERE nombre=? AND comercial_id=? AND borrado=0",
                             (datos["nombre"], comercial_id)).fetchone()
        if existe:
            saltados += 1
            continue
        campos = list(db.CAMPOS_CLIENTE)
        con.execute(f"INSERT INTO clientes ({','.join(campos)}, creado_en, creado_por, actualizado_en) "
                    f"VALUES ({','.join('?' * len(campos))},?,?,?)",
                    [datos.get(c) for c in campos] + [db.ahora(), autor["id"] if autor else None, db.ahora()])
        insertados += 1
    db.registrar(con, autor, "importar", "cliente", None, f"{insertados} clientes importados",
                 {"insertados": insertados, "duplicados_saltados": saltados}, ip)
    return insertados, saltados

@rutas.route("/admin/borme", ("GET", "POST"))
@requiere_admin
def vista_importar_borme(peticion):
    con = peticion.con
    if peticion.method == "GET":
        from datetime import date, timedelta
        ayer = date.today() - timedelta(days=1)
        while ayer.weekday() >= 5:
            ayer -= timedelta(days=1)
        fecha_defecto = ayer.isoformat()
        return render(peticion, "admin_borme.html", comerciales=comerciales(con),
                      fecha_defecto=fecha_defecto, resultado=None)

    desde_str = peticion.get("desde")
    hasta_str = peticion.get("hasta")
    destino = int(peticion.get("comercial_id") or peticion.usuario["id"])

    if not desde_str:
        return render(peticion, "admin_borme.html", comerciales=comerciales(con),
                      resultado={"error": "La fecha inicial es obligatoria."})

    from datetime import date
    try:
        desde_dt = date.fromisoformat(desde_str)
        hasta_dt = date.fromisoformat(hasta_str) if hasta_str else date.today()
    except ValueError:
        return render(peticion, "admin_borme.html", comerciales=comerciales(con),
                      resultado={"error": "Formato de fecha inválido. Utilice AAAA-MM-DD."})

    if desde_dt > hasta_dt:
        return render(peticion, "admin_borme.html", comerciales=comerciales(con),
                      resultado={"error": "La fecha de inicio no puede ser posterior a la fecha de fin."})

    prov_raw = peticion.get("provincia", "").strip()
    if prov_raw:
        filtro_provincias = {normalizar_provincia_nombre(p) for p in prov_raw.split(",") if p.strip()}
    else:
        filtro_provincias = None

    solo_constituciones = peticion.get("solo_constituciones") == "1"
    incluir_personas_fisicas = peticion.get("incluir_personas_fisicas") == "1"
    incluir_descartes = peticion.get("incluir_descartes") == "1"

    try:
        min_lineas = int(peticion.get("min_lineas") or 0)
    except ValueError:
        min_lineas = 0

    try:
        insertados, saltados, mensaje_red = descargar_e_importar_borme(
            con, desde_dt, hasta_dt, destino, filtro_provincias,
            solo_constituciones, min_lineas, incluir_personas_fisicas, incluir_descartes,
            peticion.usuario, peticion.ip
        )
        con.commit()
    except Exception as exc:  # noqa: BLE001
        traceback.print_exc()
        return render(peticion, "admin_borme.html", comerciales=comerciales(con),
                      resultado={"error": f"Error al descargar leads del BORME: {exc}"})

    return render(peticion, "admin_borme.html", comerciales=comerciales(con),
                  resultado={"insertados": insertados, "saltados": saltados, "mensaje_red": mensaje_red})


def normalizar_provincia_nombre(texto):
    import unicodedata
    if not texto:
        return ""
    sin_tildes = "".join(
        c for c in unicodedata.normalize("NFD", texto)
        if unicodedata.category(c) != "Mn"
    )
    return sin_tildes.strip().lower()


def descargar_e_importar_borme(con, desde_dt, hasta_dt, comercial_id, filtro_provincias,
                              solo_constituciones, min_lineas, incluir_personas_fisicas,
                              incluir_descartes, autor, ip="-"):
    import sys
    import os

    ruta_raiz = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    if ruta_raiz not in sys.path:
        sys.path.insert(0, ruta_raiz)

    try:
        import borme_leads
    except ImportError:
        raise Exception("No se pudo cargar el módulo borme_leads")

    cache_dir = os.path.join(ruta_raiz, "cache_borme")
    cliente = borme_leads.ClienteBORME(cache_dir=cache_dir)

    dias = list(borme_leads.rango_fechas(desde_dt, hasta_dt))
    if not dias:
        return 0, 0, " (No hay días hábiles de publicación en ese rango de fechas)"

    todas_empresas = []
    fallos_red = 0

    for dia in dias:
        try:
            items = cliente.sumario(dia)
        except borme_leads.ErrorRed:
            fallos_red += 1
            continue
        if not items:
            continue

        if filtro_provincias:
            items = [it for it in items if normalizar_provincia_nombre(it["provincia"]) in filtro_provincias]

        for it in items:
            try:
                parrafos = cliente.documento(it)
            except borme_leads.ErrorRed:
                fallos_red += 1
                continue
            if not parrafos:
                continue
            empresas = borme_leads.parsear_documento(
                parrafos, it["provincia"], it["identificador"], dia.isoformat()
            )
            todas_empresas.extend(empresas)

    for emp in todas_empresas:
        borme_leads.estimar_lineas(emp)
        borme_leads.calcular_prioridad(emp)

    filtradas = todas_empresas
    if not incluir_descartes:
        filtradas = [e for e in filtradas if not e.es_descarte]
    if not incluir_personas_fisicas:
        filtradas = [e for e in filtradas if not e.posible_persona_fisica]
    if solo_constituciones:
        filtradas = [e for e in filtradas if e.es_constitucion]
    if min_lineas:
        filtradas = [e for e in filtradas if e.lineas_estimadas >= min_lineas]

    filtradas.sort(key=lambda e: (-e.prioridad, -e.lineas_estimadas))

    insertados = 0
    saltados = 0
    for emp in filtradas:
        existe = con.execute("SELECT 1 FROM clientes WHERE nombre=? AND comercial_id=? AND borrado=0",
                             (emp.denominacion, comercial_id)).fetchone()
        if existe:
            saltados += 1
            continue

        obs_partes = []
        if emp.senal_comercial:
            obs_partes.append(f"Señal comercial: {emp.senal_comercial}")
        if emp.objeto_social:
            obs_partes.append(f"Objeto social: {emp.objeto_social}")
        if emp.domicilio:
            obs_partes.append(f"Domicilio: {emp.domicilio}")
        if emp.municipio or emp.provincia:
            obs_partes.append(f"Población: {emp.municipio} ({emp.provincia})")
        if emp.capital_eur:
            val_cap = f"{emp.capital_eur:,.2f} €".replace(",", "@").replace(".", ",").replace("@", ".")
            obs_partes.append(f"Capital social: {val_cap}")
        if emp.administradores:
            obs_partes.append(f"Administradores: {emp.administradores}")
        if emp.sector:
            obs_partes.append(f"Sector estimado: {emp.sector.replace('_', ' ').title()} (Ratio de líneas: {emp.ratio_lineas})")
        if emp.confianza:
            obs_partes.append(f"Confianza de estimación: {emp.confianza}")

        obs_texto = "\n\n".join(obs_partes)

        con.execute(
            """INSERT INTO clientes 
               (nombre, cif, persona, telefono, email, operador, producto, num_lineas, 
                cuota_linea, fecha_alta, permanencia_meses, penalizacion_total, 
                estado, proxima_accion, observaciones, comercial_id, creado_en, creado_por, actualizado_en)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                emp.denominacion,
                "", # CIF
                "", # Persona
                "", # Teléfono
                "", # Email
                "", # Operador
                "", # Producto
                emp.lineas_estimadas,
                0.0, # Cuota línea
                emp.fecha_borme,
                0, # Permanencia meses
                0.0, # Penalización total
                "Activo",
                None, # Próxima acción
                obs_texto,
                comercial_id,
                db.ahora(),
                autor["id"] if autor else None,
                db.ahora()
            )
        )
        insertados += 1

    db.registrar(con, autor, "importar_borme", "cliente", None,
                 f"{insertados} clientes importados desde BORME ({desde_dt} a {hasta_dt})",
                 {"insertados": insertados, "duplicados_saltados": saltados, "fecha_inicio": desde_dt.isoformat(), "fecha_fin": hasta_dt.isoformat()},
                 ip)

    mensaje_red = ""
    if fallos_red > 0:
        mensaje_red = f" (Ojo: fallaron {fallos_red} descargas de red)"

    return insertados, saltados, mensaje_red


@rutas.route("/exportar.xlsx")
@requiere_login
def vista_exportar(peticion):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    con = peticion.con
    lista = clientes_visibles(con, peticion.usuario)
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Cartera"
    cabeceras = ["Cliente", "CIF", "Contacto", "Teléfono", "Email", "Operador", "Producto",
                 "Nº líneas", "Cuota/línea", "Cuota total", "Fecha alta", "Permanencia",
                 "Fin permanencia", "Días restantes", "Penalización total", "Penalización pendiente",
                 "Estado", "Aviso", "Prioridad", "Comercial",
                 "Próxima acción", "Observaciones"]
    for i, texto in enumerate(cabeceras, start=1):
        celda = hoja.cell(row=1, column=i, value=texto)
        celda.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="0E2A47")
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        hoja.column_dimensions[celda.column_letter].width = 16
    for f, c in enumerate(lista, start=2):
        valores = [c["nombre"], c["cif"], c["persona"], c["telefono"], c["email"], c["operador"],
                   c["producto"], c["num_lineas"], c["cuota_linea"], c["cuota_total"],
                   negocio.a_fecha(c["fecha_alta"]), c["permanencia_meses"], c["fin_permanencia"],
                   c["dias_restantes"], c["penalizacion_total"], c["penalizacion_pendiente"],
                   c["estado"], c["aviso_texto"],
                   c["prioridad"], c.get("comercial"), negocio.a_fecha(c["proxima_accion"]),
                   c["observaciones"]]
        for i, valor in enumerate(valores, start=1):
            celda = hoja.cell(row=f, column=i, value=valor)
            celda.font = Font(name="Arial", size=9)
            if i in (9, 10, 15, 16):
                celda.number_format = '#,##0.00" €"'
            if i in (11, 13, 21):
                celda.number_format = "DD/MM/YYYY"
    hoja.freeze_panes = "A2"
    hoja.auto_filter.ref = f"A1:W{max(1, len(lista) + 1)}"
    buffer = io.BytesIO()
    libro.save(buffer)
    db.registrar(con, peticion.usuario, "exportar", "cliente", None, f"{len(lista)} clientes", None, peticion.ip)
    con.commit()
    nombre = f"cartera_{date.today().isoformat()}.xlsx"
    return Response(buffer.getvalue(), 200,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    [("Content-Disposition", f'attachment; filename="{nombre}"')])


@rutas.route("/salud")
def vista_salud(peticion):
    return json_response({"estado": "ok", "momento": db.ahora()})


# ---------------------------------------------------------------------- WSGI
def aplicacion(environ, iniciar_respuesta):
    peticion = Request(environ)
    peticion.guardar_sesion = False

    if peticion.path.startswith("/static/"):
        respuesta = servir_estatico(os.path.join(BASE, "static"), peticion.path)
        iniciar_respuesta(STATUS[respuesta.estado], respuesta.cabeceras)
        return [respuesta.cuerpo]

    con = db.conectar()
    peticion.con = con
    try:
        peticion.session = sesiones.load(peticion.cookies.get(sesiones.name))
        peticion.usuario = usuario_actual(con, peticion.session)
        peticion.ajustes = db.leer_ajustes(con)

        funcion, extra = rutas.resolver(peticion.method, peticion.path)
        if funcion is None:
            respuesta = Response(jinja.get_template("error.html").render(
                codigo=404, mensaje="Esta página no existe.", usuario=peticion.usuario,
                ruta=peticion.path, ajustes=peticion.ajustes), 404)
        else:
            # protección CSRF en cualquier envío de formulario
            if peticion.method == "POST" and peticion.path != "/login":
                esperado = peticion.session.get("csrf")
                if not esperado or peticion.form.get("csrf") != esperado:
                    respuesta = redirect(peticion.path, "La sesión ha caducado, inténtalo de nuevo.", "error")
                else:
                    respuesta = funcion(peticion, **extra)
            else:
                respuesta = funcion(peticion, **extra)

        if isinstance(respuesta, tuple):
            respuesta = respuesta[0]
        if peticion.guardar_sesion:
            if peticion.session:
                respuesta.cookie(sesiones.name, sesiones.dump(peticion.session), sesiones.max_age)
            else:
                respuesta.cookie(sesiones.name, "", borrar=True)
    except Exception:  # noqa: BLE001
        traceback.print_exc()
        respuesta = Response(jinja.get_template("error.html").render(
            codigo=500, mensaje="Se ha producido un error inesperado.", usuario=None,
            ruta=peticion.path, ajustes={}), 500)
    finally:
        con.close()

    iniciar_respuesta(STATUS.get(respuesta.estado, "200 OK"), respuesta.cabeceras)
    return [respuesta.cuerpo]


db.inicializar()
